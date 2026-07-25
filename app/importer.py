"""Excel import for the unified "one column layout per sheet" workbook format.

Every sheet in the workbook is treated as a brand, named after the sheet itself,
except the two known non-brand sheets ("ALL", a literal duplicate of every other
sheet's rows, and "STAGES", a rank-ordered reference list of stage names) - see
NON_BRAND_SHEETS. Brands, sub-brands, and stages are never hardcoded: anything
the workbook references gets created automatically if it doesn't already exist.
"""
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone

import openpyxl

from app import db

NON_BRAND_SHEETS = {"ALL", "STAGES"}

DISPATCHED_STAGE_NAME = "Dispatched"

# 1-based column indices, one layout shared by every brand sheet, data from row 2.
COLUMNS = {
    "sub_brand": 2,
    "ct": 3,
    "code": 4,
    "fabric": 5,
    "wash": 6,
    "qty": 7,
    "remark": 8,
    "fi_date": 9,
    "fabric_date": 10,
    "stage": 11,
    "days": 12,
}


@dataclass
class ImportResult:
    total_lots: int = 0
    total_pieces: int = 0
    per_brand: dict[str, int] = field(default_factory=dict)
    skipped_duplicates: list[tuple[str, str]] = field(default_factory=list)  # (sheet, ct)
    created_brands: list[str] = field(default_factory=list)
    created_sub_brands: list[tuple[str, str]] = field(default_factory=list)  # (brand, sub_brand)
    created_stages: list[str] = field(default_factory=list)
    replaced: bool = False


def _cell(row: tuple, col: int | None):
    if col is None or col - 1 >= len(row):
        return None
    return row[col - 1]


def _trim(value) -> str:
    return str(value).strip() if value is not None else ""


def _as_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _get_or_create_brand(conn: sqlite3.Connection, name: str, cache: dict, created: list) -> int:
    if name in cache:
        return cache[name]
    row = conn.execute("SELECT id FROM brands WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row is None:
        cur = conn.execute("INSERT INTO brands (name) VALUES (?)", (name,))
        brand_id = cur.lastrowid
        created.append(name)
    else:
        brand_id = row["id"]
    cache[name] = brand_id
    return brand_id


def _get_or_create_sub_brand(
    conn: sqlite3.Connection, brand_id: int, brand_name: str, name: str, cache: dict, created: list
) -> int | None:
    if not name:
        return None
    key = (brand_id, name.upper())
    if key in cache:
        return cache[key]
    row = conn.execute(
        "SELECT id FROM sub_brands WHERE brand_id = ? AND name = ? COLLATE NOCASE", (brand_id, name)
    ).fetchone()
    if row is None:
        cur = conn.execute("INSERT INTO sub_brands (brand_id, name) VALUES (?, ?)", (brand_id, name))
        sub_brand_id = cur.lastrowid
        created.append((brand_name, name))
    else:
        sub_brand_id = row["id"]
    cache[key] = sub_brand_id
    return sub_brand_id


def _get_or_create_stage(conn: sqlite3.Connection, name: str, cache: dict, created: list) -> int:
    key = name.upper()
    if key in cache:
        return cache[key]
    row = conn.execute("SELECT id FROM stages WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row is None:
        (max_rank,) = conn.execute("SELECT COALESCE(MAX(rank), 0) FROM stages").fetchone()
        cur = conn.execute("INSERT INTO stages (name, rank) VALUES (?, ?)", (name, max_rank + 1))
        stage_id = cur.lastrowid
        created.append(name)
    else:
        stage_id = row["id"]
    cache[key] = stage_id
    return stage_id


def _sync_stages_sheet(conn: sqlite3.Connection, wb, cache: dict, created: list) -> None:
    """The STAGES sheet, when present, is the authoritative rank order: any stage
    it names that already exists gets its rank updated to match; names with no
    existing match are created at that rank. Run once, before any brand rows -
    later, _get_or_create_stage covers anything actually used in the data that
    didn't exact-match a STAGES entry (e.g. a typo'd name in that sheet)."""
    if "STAGES" not in wb.sheetnames:
        return
    ws = wb["STAGES"]
    for row in ws.iter_rows(min_row=1, values_only=True):
        rank, name = _cell(row, 1), _cell(row, 2)
        if not isinstance(rank, (int, float)):
            continue
        name = _trim(name)
        if not name:
            continue
        existing = conn.execute("SELECT id FROM stages WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
        if existing is not None:
            conn.execute("UPDATE stages SET rank = ? WHERE id = ?", (int(rank), existing["id"]))
            cache[name.upper()] = existing["id"]
        else:
            cur = conn.execute("INSERT INTO stages (name, rank) VALUES (?, ?)", (name, int(rank)))
            cache[name.upper()] = cur.lastrowid
            created.append(name)


def import_workbook(
    path: str,
    conn: sqlite3.Connection,
    *,
    wipe_existing: bool = False,
    moved_by: int | None = None,
) -> ImportResult:
    """Import every brand sheet (every sheet except "ALL"/"STAGES") from the
    workbook at `path`. Brands, sub-brands, and stages referenced by the sheet
    that don't already exist are created automatically - never hardcoded.

    Rows whose CT number already exists in the app are always skipped and
    listed in `skipped_duplicates` - never silently duplicated or overwritten.
    A row whose STAGE is "Dispatched" is imported already closed (archived)
    rather than landing on the main screen. Pass `wipe_existing=True` to clear
    all lots/positions/movements first (a full "Replace" import); the default
    is an additive "Add" import.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        result = ImportResult()
        now = datetime.now(timezone.utc)
        created_at = now.isoformat()

        with db.transaction(conn):
            if wipe_existing:
                conn.execute("DELETE FROM lots")
                result.replaced = True

            brand_cache: dict[str, int] = {}
            sub_brand_cache: dict[tuple[int, str], int] = {}
            stage_cache: dict[str, int] = {}

            _sync_stages_sheet(conn, wb, stage_cache, result.created_stages)

            existing_cts = {row["ct_number"] for row in conn.execute("SELECT ct_number FROM lots")}

            for sheet_name in wb.sheetnames:
                if sheet_name in NON_BRAND_SHEETS:
                    continue
                ws = wb[sheet_name]
                brand_id = _get_or_create_brand(conn, sheet_name, brand_cache, result.created_brands)
                sheet_pieces = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    qty = _cell(row, COLUMNS["qty"])
                    if not isinstance(qty, (int, float)):
                        continue

                    stage_name = _trim(_cell(row, COLUMNS["stage"]))
                    if not stage_name:
                        continue

                    ct = _cell(row, COLUMNS["ct"])
                    ct_number = str(ct) if ct is not None else ""

                    # Some rows genuinely have a blank CT - never treat blanks as
                    # duplicates of each other.
                    if ct_number and ct_number in existing_cts:
                        result.skipped_duplicates.append((sheet_name, ct_number))
                        continue

                    stage_id = _get_or_create_stage(conn, stage_name, stage_cache, result.created_stages)

                    sub_brand_name = _trim(_cell(row, COLUMNS["sub_brand"]))
                    sub_brand_id = _get_or_create_sub_brand(
                        conn, brand_id, sheet_name, sub_brand_name, sub_brand_cache, result.created_sub_brands
                    )

                    # DAYS, when a real number, back-dates entered_at so "Days in
                    # stage" is accurate immediately - not "how long ago this lot
                    # was created", so it's used for position/movement timestamps
                    # only, never for lots.created_at (which stays the real import
                    # time - cycle-time analytics needs that to mean what it says).
                    days_value = _cell(row, COLUMNS["days"])
                    if isinstance(days_value, (int, float)) and days_value >= 0:
                        entered_at = (now - timedelta(days=days_value)).isoformat()
                    else:
                        entered_at = created_at

                    wash = _cell(row, COLUMNS["wash"])
                    remark = _cell(row, COLUMNS["remark"])

                    cur = conn.execute(
                        "INSERT INTO lots "
                        "(brand_id, sub_brand_id, ct_number, material_code, fabric, wash, "
                        "total_qty, fi_date, fabric_date, remark, created_at) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            brand_id,
                            sub_brand_id,
                            ct_number,
                            str(_cell(row, COLUMNS["code"]) or "") or None,
                            str(_cell(row, COLUMNS["fabric"]) or "") or None,
                            str(wash) if wash else None,
                            int(qty),
                            _as_date(_cell(row, COLUMNS["fi_date"])),
                            _as_date(_cell(row, COLUMNS["fabric_date"])),
                            str(remark) if remark else None,
                            created_at,
                        ),
                    )
                    lot_id = cur.lastrowid
                    conn.execute(
                        "INSERT INTO positions (lot_id, stage_id, qty, entered_at) VALUES (?, ?, ?, ?)",
                        (lot_id, stage_id, int(qty), entered_at),
                    )
                    conn.execute(
                        "INSERT INTO movements "
                        "(lot_id, from_stage_id, to_stage_id, qty, moved_at, moved_by, note) "
                        "VALUES (?, NULL, ?, ?, ?, ?, ?)",
                        (lot_id, stage_id, int(qty), entered_at, moved_by, "opening import"),
                    )

                    if stage_name.upper() == DISPATCHED_STAGE_NAME.upper():
                        conn.execute("UPDATE lots SET closed_at = ? WHERE id = ?", (entered_at, lot_id))

                    if ct_number:
                        existing_cts.add(ct_number)
                    result.total_lots += 1
                    result.total_pieces += int(qty)
                    sheet_pieces += int(qty)

                result.per_brand[sheet_name] = sheet_pieces

        return result
    finally:
        wb.close()
