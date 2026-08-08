"""Export .xlsx column layout: no "Wash", "Order Qty" instead of "Qty", and a
new "ACR" column with the same split-lot dedupe rule as fetch_grouped_positions
(see main_screen.py's comment - ACR is a flat per-lot value, so a split lot's
ACR must count once in the true grand total even though it appears on more
than one detail row)."""

import openpyxl

from app import operations
from app.exporter import export_workbook


def _create_lot(conn, brand_ids, stage_ids, user_id, qty=1000, ct="A100"):
    return operations.create_lot(
        conn,
        brand_id=brand_ids["SPYKAR"],
        ct_number=ct,
        total_qty=qty,
        starting_stage_id=stage_ids["Under Cutting"],
        moved_by=user_id,
    )


def _load(buffer):
    return openpyxl.load_workbook(buffer)


def test_header_row_has_no_wash_and_renamed_qty_columns(conn, brand_ids, stage_ids, user_id):
    _create_lot(conn, brand_ids, stage_ids, user_id)
    wb = _load(export_workbook(conn))
    header = [cell.value for cell in next(wb["DPR"].iter_rows(min_row=1, max_row=1))]
    assert "Wash" not in header
    assert "Order Qty" in header
    assert "ACR" in header
    assert "Qty" not in header


def test_detail_row_includes_acr_value(conn, brand_ids, stage_ids, user_id):
    lot_id = _create_lot(conn, brand_ids, stage_ids, user_id, qty=500)
    operations.update_lot_details(conn, lot_id=lot_id, acr=400)

    wb = _load(export_workbook(conn))
    rows = list(wb["DPR"].iter_rows(min_row=1, values_only=True))
    header = rows[0]
    acr_col = header.index("ACR")
    ct_col = header.index("CT")
    detail_row = next(r for r in rows[1:] if r[ct_col] == "A100")
    assert detail_row[acr_col] == 400


def test_grand_total_dedupes_a_split_lots_acr(conn, brand_ids, stage_ids, user_id):
    lot_id = _create_lot(conn, brand_ids, stage_ids, user_id, qty=1000)
    operations.update_lot_details(conn, lot_id=lot_id, acr=600)
    operations.move_pieces(
        conn, lot_id=lot_id, from_stage_id=stage_ids["Under Cutting"],
        to_stage_id=stage_ids["Under Sewing"], qty=400, moved_by=user_id,
    )

    wb = _load(export_workbook(conn))
    detail_rows = list(wb["DPR"].iter_rows(min_row=1, values_only=True))
    header = detail_rows[0]
    label_col = header.index("Fabric")  # "Subtotal"/"Grand total" label sits here
    grand_total_row = next(r for r in detail_rows if r[label_col] == "Grand total")
    assert grand_total_row[label_col + 2] == 600  # ACR grand total, not 1200

    report_rows = list(wb["Report"].iter_rows(min_row=1, values_only=True))
    report_header = report_rows[0]
    report_grand_total = next(r for r in report_rows if r[0] == "Grand total")
    assert report_grand_total[report_header.index("ACR")] == 600
