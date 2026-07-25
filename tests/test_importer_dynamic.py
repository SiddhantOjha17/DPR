"""Unit tests for the dynamic (brand/sub-brand/stage auto-creation) importer
behavior, using small synthetic workbooks built in-test rather than the big real
fixture, so each behavior is isolated and easy to reason about. The real-file
based tests (test_import_routes.py, test_timezone_consistency.py) cover the
end-to-end reconciliation case."""
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest

from app.importer import import_workbook
from app.invariants import check_invariants

HEADER = ["BRAND", "SUB", "CT", "STYLE", "Fabric", "WASH", "Qty", "remark", "FI DATE", "FABRIC DATE", "STAGE", "DAYS"]


def _make_workbook(tmp_path, sheets: dict[str, list[list]], stages_rows: list[tuple] | None = None):
    """sheets: {sheet_name: [row, row, ...]} (header added automatically for
    non-STAGES sheets). stages_rows: list of (rank, name) tuples for a STAGES sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(HEADER)
        for row in rows:
            ws.append(row)
    if stages_rows is not None:
        ws = wb.create_sheet("STAGES")
        ws.append(["RANK", "STAGE"])
        for rank, name in stages_rows:
            ws.append([rank, name])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return str(path)


def _row(ct, sub="", qty=100, stage="Under Cutting", code="", fabric="", wash="", fi_date=None, fabric_date=None, remark="", days=None):
    return ["BRANDNAME", sub, ct, code, fabric, wash, qty, remark, fi_date, fabric_date, stage, days]


def test_new_brand_auto_created_from_sheet_name(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1")]})
    result = import_workbook(path, conn)
    assert result.created_brands == ["NEWBRAND"]
    brand = conn.execute("SELECT * FROM brands WHERE name = 'NEWBRAND'").fetchone()
    assert brand is not None
    lot = conn.execute("SELECT * FROM lots WHERE ct_number = 'C1'").fetchone()
    assert lot["brand_id"] == brand["id"]


def test_existing_brand_not_recreated(tmp_path, conn, brand_ids):
    path = _make_workbook(tmp_path, {"SPYKAR": [_row("C1")]})
    result = import_workbook(path, conn)
    assert result.created_brands == []
    lot = conn.execute("SELECT * FROM lots WHERE ct_number = 'C1'").fetchone()
    assert lot["brand_id"] == brand_ids["SPYKAR"]


def test_sub_brand_auto_created_for_any_brand(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1", sub="Some Sub Line")]})
    result = import_workbook(path, conn)
    assert result.created_sub_brands == [("NEWBRAND", "Some Sub Line")]
    row = conn.execute(
        "SELECT sb.name FROM lots l JOIN sub_brands sb ON sb.id = l.sub_brand_id WHERE l.ct_number = 'C1'"
    ).fetchone()
    assert row["name"] == "Some Sub Line"


def test_all_and_stages_sheets_never_treated_as_brands(tmp_path, conn):
    path = _make_workbook(
        tmp_path,
        {"NEWBRAND": [_row("C1")], "ALL": [_row("C1")]},  # ALL duplicates NEWBRAND's row
        stages_rows=[(1, "Dispatched"), (2, "Under Cutting")],
    )
    result = import_workbook(path, conn)
    assert "ALL" not in result.per_brand
    assert "STAGES" not in result.per_brand
    assert conn.execute("SELECT 1 FROM brands WHERE name = 'ALL'").fetchone() is None
    assert conn.execute("SELECT 1 FROM brands WHERE name = 'STAGES'").fetchone() is None
    # only NEWBRAND's copy of C1 imported, not a second one from "ALL"
    assert result.total_lots == 1


def test_stages_sheet_reorders_existing_stages(tmp_path, conn, stage_ids):
    # Under Cutting is normally rank 8; make the STAGES sheet say it should be rank 1.
    path = _make_workbook(
        tmp_path, {"NEWBRAND": [_row("C1", stage="Under Cutting")]},
        stages_rows=[(1, "Under Cutting")],
    )
    import_workbook(path, conn)
    row = conn.execute("SELECT rank FROM stages WHERE id = ?", (stage_ids["Under Cutting"],)).fetchone()
    assert row["rank"] == 1


def test_stages_sheet_typo_falls_back_to_real_usage(tmp_path, conn):
    # "STAGES" sheet has a typo'd name that never actually appears in the data;
    # the real (correctly spelled) name used by an actual row must still get
    # created, even though it doesn't match the typo'd sheet entry.
    path = _make_workbook(
        tmp_path,
        {"NEWBRAND": [_row("C1", stage="Under Size Set")]},
        stages_rows=[(5, "Under Size Seted")],  # typo, never matches real usage
    )
    result = import_workbook(path, conn)
    assert "Under Size Seted" in result.created_stages  # created from the sheet, unused
    assert "Under Size Set" in result.created_stages  # created from real row usage
    lot_stage = conn.execute(
        "SELECT s.name FROM lots l JOIN positions p ON p.lot_id = l.id JOIN stages s ON s.id = p.stage_id "
        "WHERE l.ct_number = 'C1'"
    ).fetchone()
    assert lot_stage["name"] == "Under Size Set"


def test_dispatched_row_imports_directly_into_archive(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1", stage="Dispatched")]})
    import_workbook(path, conn)
    lot = conn.execute("SELECT * FROM lots WHERE ct_number = 'C1'").fetchone()
    assert lot["closed_at"] is not None
    position = conn.execute("SELECT * FROM positions WHERE lot_id = ?", (lot["id"],)).fetchone()
    stage = conn.execute("SELECT name FROM stages WHERE id = ?", (position["stage_id"],)).fetchone()
    assert stage["name"] == "Dispatched"


def test_non_dispatched_row_stays_open(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1", stage="Under Cutting")]})
    import_workbook(path, conn)
    lot = conn.execute("SELECT closed_at FROM lots WHERE ct_number = 'C1'").fetchone()
    assert lot["closed_at"] is None


def test_days_column_backdates_entered_at_when_numeric(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1", days=10)]})
    before = datetime.now(timezone.utc)
    import_workbook(path, conn)
    lot = conn.execute("SELECT id, created_at FROM lots WHERE ct_number = 'C1'").fetchone()
    position = conn.execute("SELECT entered_at FROM positions WHERE lot_id = ?", (lot["id"],)).fetchone()
    entered = datetime.fromisoformat(position["entered_at"])
    created = datetime.fromisoformat(lot["created_at"])
    # entered_at should be ~10 days before "now", created_at should be ~"now" -
    # these must differ (DAYS backdates position/movement, never lots.created_at,
    # since DAYS means "days in current stage", not "days since lot was opened").
    assert (before - entered).days in (9, 10, 11)
    assert abs((created - before).total_seconds()) < 5


def test_days_column_blank_falls_back_to_import_time(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1", days=None)]})
    before = datetime.now(timezone.utc)
    import_workbook(path, conn)
    lot_id = conn.execute("SELECT id FROM lots WHERE ct_number = 'C1'").fetchone()["id"]
    position = conn.execute("SELECT entered_at FROM positions WHERE lot_id = ?", (lot_id,)).fetchone()
    entered = datetime.fromisoformat(position["entered_at"])
    assert abs((entered - before).total_seconds()) < 5


def test_duplicate_ct_still_skipped_and_reported(tmp_path, conn):
    path = _make_workbook(tmp_path, {"NEWBRAND": [_row("C1"), _row("C1")]})
    result = import_workbook(path, conn)
    assert result.total_lots == 1
    assert result.skipped_duplicates == [("NEWBRAND", "C1")]


def test_invariant_holds_after_import(tmp_path, conn):
    path = _make_workbook(
        tmp_path,
        {"NEWBRAND": [_row("C1", qty=100, stage="Under Cutting"), _row("C2", qty=200, stage="Dispatched")]},
    )
    import_workbook(path, conn)
    check_invariants(conn)  # should not raise
